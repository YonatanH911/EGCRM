"""
Import script:
1. Creates the master user Ronen (if not exists)
2. Imports all accounts from the Active Accounts Excel export
"""
import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from sqlalchemy.orm import Session
from database import SessionLocal, engine
import models, auth

EXCEL_PATH = r'c:\projects\CRM_Dynamics\excel_exports\Active Accounts 26-Feb-26 1-28-48 PM.xlsx'

def run():
    db: Session = SessionLocal()
    try:
        # ------- 1. Create Ronen master user if not exists -------
        existing_user = db.query(models.User).filter(models.User.email == "ronen@crm.com").first()
        if existing_user:
            print(f"User Ronen already exists (id={existing_user.id})")
        else:
            ronen = models.User(
                email="ronen@crm.com",
                name="Ronen",
                password_hash=auth.get_password_hash("1111"),
                role=models.UserRole.ADMIN,
            )
            db.add(ronen)
            db.commit()
            db.refresh(ronen)
            print(f"Created master user Ronen (id={ronen.id})")

        # ------- 2. Import accounts -------
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        print("Headers found:", headers)

        # Map Excel columns -> model fields
        # Headers: ['(Do Not Modify) Account', '(Do Not Modify) Row Checksum', '(Do Not Modify) Modified On',
        #           'Comapny Name', 'Phone', 'Street', 'ZIP/Postal Code', 'City', 'Country/Region', 'Website']
        col_idx = {h: i for i, h in enumerate(headers)}

        imported = 0
        skipped = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            company_name = row[col_idx.get('Comapny Name', 3)]  # typo in their header
            if not company_name:
                skipped += 1
                continue

            phone = row[col_idx.get('Phone', 4)]
            street = row[col_idx.get('Street', 5)]
            zip_code = row[col_idx.get('ZIP/Postal Code', 6)]
            city = row[col_idx.get('City', 7)]
            country = row[col_idx.get('Country/Region', 8)]
            website = row[col_idx.get('Website', 9)]

            # Skip if account with exact same name already exists
            exists = db.query(models.Account).filter(models.Account.name == company_name).first()
            if exists:
                skipped += 1
                continue

            account = models.Account(
                name=str(company_name).strip() if company_name else None,
                phone=str(phone).strip() if phone else None,
                street=str(street).strip() if street else None,
                zip_code=str(zip_code).strip() if zip_code else None,
                city=str(city).strip() if city else None,
                country=str(country).strip() if country else None,
                website=str(website).strip() if website else None,
            )
            db.add(account)
            imported += 1

        db.commit()
        print(f"\n✅ Done! Imported: {imported} accounts | Skipped: {skipped}")

    except Exception as e:
        db.rollback()
        print(f"❌ Error: {e}")
        raise
    finally:
        db.close()

if __name__ == '__main__':
    run()
