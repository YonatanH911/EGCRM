"""Restore contract contact fields from an Excel export.

The script is intentionally a dry run unless --apply is supplied. Existing CRM
values are preserved by default; use --overwrite only when the spreadsheet must
replace values that are already present.

Examples:
    python migrate_contract_contacts.py "../excel_exports/Active Contracts.xlsx"
    python migrate_contract_contacts.py "../excel_exports/Active Contracts.xlsx" --apply
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Iterable

import openpyxl
from sqlalchemy.orm import Session

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import models
from database import SessionLocal


CONTACT_FIELDS = {
    "beneficiary_management_contact": ("beneficiary", "management"),
    "beneficiary_technical_contact": ("beneficiary", "technical"),
    "beneficiary_financial_contact": ("beneficiary", "financial"),
    "supplier_management_contact": ("supplier", "management"),
    "supplier_technical_contact": ("supplier", "technical"),
    "supplier_financial_contact": ("supplier", "financial"),
}

SIDE_WORDS = {
    "beneficiary": {"beneficiary", "benefeciary"},
    "supplier": {"supplier", "vendor"},
}

ROLE_WORDS = {
    "management": {"management", "manager", "managment"},
    "technical": {"technical", "tech"},
    "financial": {"financial", "finance", "billing"},
}

TITLE_HEADERS = {
    "name",
    "contract name",
    "contract title",
    "title",
}


def text_value(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_words(value: object) -> list[str]:
    normalized = text_value(value).casefold()
    normalized = normalized.replace("benefeciary", "beneficiary")
    normalized = normalized.replace("managment", "management")
    return re.findall(r"[a-z0-9]+", normalized)


def normalize_title(value: object) -> str:
    normalized = text_value(value).casefold()
    normalized = re.sub(r"[\u2010-\u2015]", "-", normalized)
    normalized = re.sub(r"\s*-\s*", " - ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def normalize_contact(value: object) -> str:
    return " ".join(normalize_words(value))


def find_title_column(headers: list[object]) -> int | None:
    normalized = [" ".join(normalize_words(header)) for header in headers]
    for alias in TITLE_HEADERS:
        if alias in normalized:
            return normalized.index(alias)
    return None


def has_any(tokens: set[str], choices: Iterable[str]) -> bool:
    return bool(tokens.intersection(choices))


def resolve_contact_columns(
    headers: list[object], legacy_side: str
) -> tuple[dict[str, int], list[str]]:
    tokenized = [set(normalize_words(header)) for header in headers]
    resolved: dict[str, int] = {}
    warnings: list[str] = []

    for field, (side, role) in CONTACT_FIELDS.items():
        candidates = [
            index
            for index, tokens in enumerate(tokenized)
            if has_any(tokens, SIDE_WORDS[side])
            and has_any(tokens, ROLE_WORDS[role])
        ]
        if candidates:
            resolved[field] = candidates[0]
            if len(candidates) > 1:
                names = ", ".join(text_value(headers[index]) for index in candidates)
                warnings.append(
                    f"Multiple columns match {field}; using '{text_value(headers[candidates[0]])}' "
                    f"from: {names}"
                )

    all_side_words = set().union(*SIDE_WORDS.values())
    for role, role_aliases in ROLE_WORDS.items():
        candidates = [
            index
            for index, tokens in enumerate(tokenized)
            if has_any(tokens, role_aliases)
            and not has_any(tokens, all_side_words)
            and ("contact" in tokens or len(tokens) <= 2)
        ]
        if not candidates or legacy_side == "ignore":
            continue

        legacy_index = candidates[0]
        target_sides = (
            ("beneficiary", "supplier") if legacy_side == "both" else (legacy_side,)
        )
        for side in target_sides:
            field = f"{side}_{role}_contact"
            if field not in resolved:
                resolved[field] = legacy_index
                warnings.append(
                    f"Using legacy column '{text_value(headers[legacy_index])}' for {field}."
                )

    return resolved, warnings


def find_default_workbook() -> Path | None:
    export_dir = Path(__file__).resolve().parent.parent / "excel_exports"
    candidates = sorted(
        export_dir.glob("*Contract*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def contract_keys(contract: models.Contract) -> set[str]:
    keys = {normalize_title(contract.title)}
    if contract.beneficiary_title and contract.supplier_title:
        keys.add(
            normalize_title(
                f"{contract.beneficiary_title} - {contract.supplier_title}"
            )
        )
    return {key for key in keys if key}


def build_contact_lookup(contacts: list[models.Contact]):
    names: dict[str, set[str]] = defaultdict(set)
    emails: dict[str, str] = {}
    for contact in contacts:
        full_name = text_value(f"{contact.first_name or ''} {contact.last_name or ''}")
        if full_name:
            names[normalize_contact(full_name)].add(full_name)
        if contact.email:
            emails[contact.email.strip().casefold()] = full_name
    return names, emails


def resolve_contact_piece(
    value: str,
    names: dict[str, set[str]],
    emails: dict[str, str],
) -> tuple[str, bool]:
    value = text_value(value)
    if not value:
        return "", True

    email_match = re.search(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}", value)
    if email_match:
        canonical = emails.get(email_match.group(0).casefold())
        if canonical:
            return canonical, True

    matches = names.get(normalize_contact(value), set())
    if len(matches) == 1:
        return next(iter(matches)), True
    return value, False


def canonicalize_contact_cell(
    value: object,
    names: dict[str, set[str]],
    emails: dict[str, str],
) -> tuple[str, list[str]]:
    raw = text_value(value)
    if not raw:
        return "", []

    pieces = [piece for piece in re.split(r"[;|\n]+", raw) if text_value(piece)]
    if len(pieces) == 1 and "," in raw:
        whole, whole_matched = resolve_contact_piece(raw, names, emails)
        comma_parts = [text_value(piece) for piece in raw.split(",") if text_value(piece)]
        resolved_parts = [resolve_contact_piece(piece, names, emails) for piece in comma_parts]
        if not whole_matched and len(comma_parts) > 1 and all(match for _, match in resolved_parts):
            pieces = comma_parts

    values: list[str] = []
    unresolved: list[str] = []
    seen: set[str] = set()
    for piece in pieces:
        canonical, matched = resolve_contact_piece(piece, names, emails)
        key = canonical.casefold()
        if not canonical or key in seen:
            continue
        seen.add(key)
        values.append(canonical)
        if not matched:
            unresolved.append(canonical)
    return ", ".join(values), unresolved


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Restore the six contract contact fields from an Excel export."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        type=Path,
        help="Contracts .xlsx file. Defaults to the newest *Contract*.xlsx in excel_exports.",
    )
    parser.add_argument("--sheet", help="Worksheet name. Defaults to the active worksheet.")
    parser.add_argument("--header-row", type=int, default=1, help="One-based header row (default: 1).")
    parser.add_argument("--apply", action="store_true", help="Commit the displayed changes.")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace populated CRM contact fields instead of filling only empty fields.",
    )
    parser.add_argument(
        "--legacy-side",
        choices=("beneficiary", "supplier", "both", "ignore"),
        default="both",
        help="Where generic legacy contact columns are restored (default: both).",
    )
    return parser.parse_args()


def run() -> int:
    args = parse_args()
    workbook_path = args.workbook or find_default_workbook()
    if workbook_path is None:
        print("No Contracts workbook found. Pass its path as the first argument.")
        return 2

    workbook_path = workbook_path.expanduser().resolve()
    if not workbook_path.is_file():
        print(f"Workbook not found: {workbook_path}")
        return 2
    if args.header_row < 1:
        print("--header-row must be at least 1.")
        return 2

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if args.sheet:
        if args.sheet not in workbook.sheetnames:
            print(f"Worksheet '{args.sheet}' not found. Available: {', '.join(workbook.sheetnames)}")
            workbook.close()
            return 2
        worksheet = workbook[args.sheet]
    else:
        worksheet = workbook.active

    header_values = next(
        worksheet.iter_rows(
            min_row=args.header_row,
            max_row=args.header_row,
            values_only=True,
        )
    )
    headers = list(header_values)
    title_index = find_title_column(headers)
    contact_columns, mapping_warnings = resolve_contact_columns(headers, args.legacy_side)

    print(f"Workbook: {workbook_path}")
    print(f"Worksheet: {worksheet.title}")
    print(f"Mode: {'APPLY' if args.apply else 'DRY RUN'}")
    if title_index is None:
        print("No contract title column found. Expected Name, Contract Name, Contract Title, or Title.")
        print("Detected headers: " + ", ".join(text_value(header) for header in headers if header))
        workbook.close()
        return 2
    if not contact_columns:
        print("No beneficiary or supplier contact columns were recognized.")
        print("Detected headers: " + ", ".join(text_value(header) for header in headers if header))
        workbook.close()
        return 2

    print(f"Contract key: {text_value(headers[title_index])}")
    for field in CONTACT_FIELDS:
        index = contact_columns.get(field)
        source = text_value(headers[index]) if index is not None else "not found"
        print(f"  {field} <- {source}")
    for warning in mapping_warnings:
        print(f"WARNING: {warning}")

    db: Session = SessionLocal()
    try:
        contracts = db.query(models.Contract).all()
        contacts = db.query(models.Contact).all()
        index: dict[str, list[models.Contract]] = defaultdict(list)
        for contract in contracts:
            for key in contract_keys(contract):
                index[key].append(contract)

        contact_names, contact_emails = build_contact_lookup(contacts)
        rows_seen = matched_rows = unmatched_rows = ambiguous_rows = 0
        changed_contracts: set[int] = set()
        changed_fields = 0
        unresolved_values: set[str] = set()

        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=args.header_row + 1, values_only=True),
            start=args.header_row + 1,
        ):
            rows_seen += 1
            title = text_value(row[title_index]) if title_index < len(row) else ""
            if not title:
                continue

            matches = index.get(normalize_title(title), [])
            unique_matches = {contract.id: contract for contract in matches}
            if not unique_matches:
                unmatched_rows += 1
                print(f"UNMATCHED row {row_number}: {title}")
                continue
            if len(unique_matches) > 1:
                ambiguous_rows += 1
                print(f"AMBIGUOUS row {row_number}: {title}")
                continue

            matched_rows += 1
            contract = next(iter(unique_matches.values()))
            row_changes: list[str] = []
            for field, column_index in contact_columns.items():
                if column_index >= len(row):
                    continue
                incoming, unresolved = canonicalize_contact_cell(
                    row[column_index], contact_names, contact_emails
                )
                unresolved_values.update(unresolved)
                if not incoming:
                    continue

                current = text_value(getattr(contract, field, None))
                if current and not args.overwrite:
                    continue
                if current == incoming:
                    continue

                setattr(contract, field, incoming)
                changed_fields += 1
                changed_contracts.add(contract.id)
                row_changes.append(f"{field}='{incoming}'")

            if row_changes:
                print(f"UPDATE contract {contract.id} ({title}): " + "; ".join(row_changes))

        if args.apply:
            db.commit()
        else:
            db.rollback()

        print("\nSummary")
        print(f"  Excel rows scanned: {rows_seen}")
        print(f"  Rows matched: {matched_rows}")
        print(f"  Rows unmatched: {unmatched_rows}")
        print(f"  Rows ambiguous: {ambiguous_rows}")
        print(f"  Contracts changed: {len(changed_contracts)}")
        print(f"  Contact fields changed: {changed_fields}")
        if unresolved_values:
            print("  Values not matched to Contacts records (preserved as written):")
            for value in sorted(unresolved_values, key=str.casefold):
                print(f"    - {value}")
        print("Changes committed." if args.apply else "Dry run only; no database changes were committed.")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()
        workbook.close()


if __name__ == "__main__":
    raise SystemExit(run())
