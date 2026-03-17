"""
Import script for My Active Contacts Excel export.
Maps: Full Name, First Name, Middle Name, Last Name, Email,
      Business Phone, Mobile Phone, Account, Israeli?
Links contacts to existing Account records by name.
"""
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from sqlalchemy.orm import Session
from database import SessionLocal
import models

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'excel_exports')

def str_val(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def import_contacts(db: Session):
    print("\n--- Importing Contacts ---")
    wb = openpyxl.load_workbook(
        os.path.join(EXPORTS_DIR, 'My Active Contacts 26-Feb-26 1-29-17 PM.xlsx'),
        read_only=True
    )
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    # Build account name -> id lookup
    account_map = {}
    for acc in db.query(models.Account).all():
        if acc.name:
            account_map[acc.name.strip().lower()] = acc.id

    imported = updated = skipped = 0
    seen_emails = set()   # handle duplicate emails within the Excel
    seen_names  = set()   # handle duplicate name-based contacts

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name = str_val(row[col.get('First Name', 4)])
        last_name  = str_val(row[col.get('Last Name', 6)])
        if not first_name and not last_name:
            skipped += 1
            continue

        first_name = first_name or ''
        last_name  = last_name  or ''

        email       = str_val(row[col.get('Email', 7)])
        middle_name = str_val(row[col.get('Middle Name', 5)])
        phone       = str_val(row[col.get('Business Phone', 9)])
        mobile      = str_val(row[col.get('Mobile Phone', 10)])
        account_raw = str_val(row[col.get('Account', 8)])
        israeli_raw = str_val(row[col.get('Israeli?', 11)])
        is_israeli  = True if israeli_raw and israeli_raw.lower() == 'yes' else False

        account_id = None
        if account_raw:
            account_id = account_map.get(account_raw.strip().lower())

        # Deduplicate by email (within Excel and across DB)
        if email:
            email_lower = email.lower()
            if email_lower in seen_emails:
                skipped += 1
                continue
            seen_emails.add(email_lower)

            existing = db.query(models.Contact).filter(
                models.Contact.email == email
            ).first()
            if existing:
                existing.middle_name  = middle_name
                existing.mobile_phone = mobile
                existing.is_israeli   = is_israeli
                if account_id and not existing.account_id:
                    existing.account_id = account_id
                db.commit()
                updated += 1
                continue
        else:
            # No email — deduplicate by first+last name
            name_key = (first_name.lower(), last_name.lower())
            if name_key in seen_names:
                skipped += 1
                continue
            seen_names.add(name_key)

            existing = db.query(models.Contact).filter(
                models.Contact.first_name == first_name,
                models.Contact.last_name  == last_name,
                models.Contact.email      == None,
            ).first()
            if existing:
                existing.middle_name  = middle_name
                existing.mobile_phone = mobile
                existing.is_israeli   = is_israeli
                db.commit()
                updated += 1
                continue

        contact = models.Contact(
            first_name  = first_name,
            last_name   = last_name,
            middle_name = middle_name,
            email       = email,
            phone       = phone,
            mobile_phone= mobile,
            is_israeli  = is_israeli,
            account_id  = account_id,
        )
        db.add(contact)
        db.commit()
        imported += 1

    wb.close()
    print(f"  Contacts: imported={imported}, updated={updated}, skipped={skipped}")


def run():
    db: Session = SessionLocal()
    try:
        import_contacts(db)
        print("\nContact import complete.")
    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        import traceback; traceback.print_exc()
        raise
    finally:
        db.close()


if __name__ == '__main__':
    run()
