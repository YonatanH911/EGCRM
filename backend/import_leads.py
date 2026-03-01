"""
Import script: My Open Leads Excel -> leads table
Maps: Topic -> title, Status Reason -> status, First/Last Name used to match contact_id
"""
import os
import sys
import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session
from dotenv import load_dotenv

load_dotenv()

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r'c:\projects\CRM_Dynamics\excel_exports\My Open Leads 26-Feb-26 4-36-09 PM.xlsx'

# Map Dynamics status values -> our LeadStatus enum values
STATUS_MAP = {
    'New': 'New',
    'Contacted': 'Contacted',
    'Qualified': 'Qualified',
    'Lost': 'Lost',
    'Cannot Contact': 'Lost',
    'No Longer Interested': 'Lost',
    'Canceled': 'Lost',
}

DATABASE_URL = os.getenv("DATABASE_URL", "mysql+pymysql://root@localhost:3306/crm_db")
engine = create_engine(DATABASE_URL)

wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb['My Open Leads']

headers = [c.value for c in ws[1]]
print(f"Headers: {headers}")

# Col indices (0-based)
IDX_FIRST   = headers.index('First Name')
IDX_MIDDLE  = headers.index('Middle Name')
IDX_LAST    = headers.index('Last Name')
IDX_TOPIC   = headers.index('Topic')
IDX_STATUS  = headers.index('Status Reason')

imported = 0
skipped  = 0

with Session(engine) as session:
    # Load existing lead titles to avoid exact duplicates
    existing_titles = {row[0].lower() for row in session.execute(text("SELECT title FROM leads")).fetchall()}

    # Load contacts for fuzzy matching: (first_lower, last_lower) -> contact_id
    contacts_raw = session.execute(
        text("SELECT id, LOWER(first_name), LOWER(last_name) FROM contacts")
    ).fetchall()
    contact_map = {(r[1].strip(), r[2].strip()): r[0] for r in contacts_raw}

    for row in ws.iter_rows(min_row=2, values_only=True):
        first   = (row[IDX_FIRST]  or '').strip()
        middle  = (row[IDX_MIDDLE] or '').strip()
        last    = (row[IDX_LAST]   or '').strip()
        topic   = (row[IDX_TOPIC]  or '').strip()
        status_raw = (row[IDX_STATUS] or 'New').strip()

        title = topic if topic else f"{first} {last}".strip()

        if not title:
            print(f"  SKIP: no title or name")
            skipped += 1
            continue

        if title.lower() in existing_titles:
            print(f"  SKIP (duplicate): {title!r}")
            skipped += 1
            continue

        status = STATUS_MAP.get(status_raw, 'New')

        # Try to match contact by first+last name
        contact_id = contact_map.get((first.lower(), last.lower()))
        if contact_id is None:
            # Try last-name only if last contains spaces (e.g. "cohen hav ron")
            for (cf, cl), cid in contact_map.items():
                if cl == last.lower() and (not first or cf == first.lower()):
                    contact_id = cid
                    break

        session.execute(
            text("""
                INSERT INTO leads (title, status, value, contact_id, assigned_to_user_id, created_at)
                VALUES (:title, :status, 0.0, :contact_id, NULL, NOW())
            """),
            {"title": title, "status": status, "contact_id": contact_id}
        )
        existing_titles.add(title.lower())
        match_note = f"→ contact {contact_id}" if contact_id else "→ no contact match"
        print(f"  IMPORTED [{status}]: {title!r}  {match_note}")
        imported += 1

    session.commit()

print(f"\nDone. Imported: {imported}, Skipped: {skipped}")
