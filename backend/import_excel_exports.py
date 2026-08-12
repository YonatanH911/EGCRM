"""Import active and inactive Dynamics Excel exports into EGCRM.

The importer is a dry run by default. It uses the Dynamics GUID in each export
to make repeated runs stable and imports records in dependency order.

Usage:
    python import_excel_exports.py ../excel
    python import_excel_exports.py ../excel --apply
    python import_excel_exports.py ../excel --apply --overwrite
"""
from __future__ import annotations

import argparse
import re
import sys
import warnings
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import text
from sqlalchemy.orm import Session

from database import SessionLocal
import models


warnings.filterwarnings("ignore", message="Workbook contains no default style")

FILES = {
    "active_accounts": "Active Accounts .xlsx",
    "inactive_accounts": "Inactive Accounts.xlsx",
    "active_contacts": "Active Contacts.xlsx",
    "inactive_contacts": "Inactive Contacts.xlsx",
    "active_contracts": "Active Contracts.xlsx",
    "inactive_contracts": "Inactive Contracts.xlsx",
    "active_deposits": "Active Deposits.xlsx",
    "inactive_deposits": "Inactive Deposits.xlsx",
    "activities": "Activities.xlsx",
}


@dataclass
class Stats:
    scanned: int = 0
    created: int = 0
    matched: int = 0
    changed: int = 0
    skipped: int = 0
    unresolved: int = 0

    def display(self, label: str) -> None:
        print(
            f"{label}: scanned={self.scanned}, created={self.created}, "
            f"matched={self.matched}, changed={self.changed}, "
            f"skipped={self.skipped}, unresolved={self.unresolved}"
        )


def clean(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalized(value: Any) -> str:
    return clean(value).casefold()


def normalized_phone(value: Any) -> str:
    return re.sub(r"\D", "", clean(value))


def normalize_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", normalized(value)).strip()


def parse_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    raw = clean(value)
    for pattern in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, pattern)
        except ValueError:
            continue
    return None


def parse_yes_no(value: Any) -> bool | None:
    raw = normalized(value)
    if raw in {"yes", "true", "1"}:
        return True
    if raw in {"no", "false", "0"}:
        return False
    return None


def split_name(first: Any, middle: Any, last: Any, full: Any) -> tuple[str, str, str]:
    first_name = clean(first)
    middle_name = clean(middle)
    last_name = clean(last)
    if first_name or last_name:
        return first_name, middle_name, last_name
    parts = clean(full).split()
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def split_contract_title(title: str) -> tuple[str, str]:
    parts = re.split(r"\s+-\s+", clean(title), maxsplit=1)
    if len(parts) == 1:
        return parts[0], ""
    return parts[0].strip(), parts[1].strip()


def activity_type_from_subject(subject: Any) -> str:
    """Extract the activity category written at the start of an imported subject."""
    value = clean(subject)
    lowered = value.casefold()
    known_prefixes = (
        (r"^billing\b", "Billing"),
        (r"^dep(?:osit)?\s+req(?:uest)?\b", "Dep Req"),
        (r"^ronen\b", "Ronen"),
        (r"^termination\b", "Termination"),
    )
    for pattern, activity_type in known_prefixes:
        if re.match(pattern, lowered):
            return activity_type

    parts = re.split(r"\s+-\s+", value, maxsplit=1)
    return parts[0].strip() if len(parts) == 2 else ""


class WorkbookRows:
    def __init__(self, path: Path):
        self.path = path
        self.workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        self.worksheet = self.workbook.active
        raw_headers = next(self.worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        self.columns = {
            normalize_header(header): index
            for index, header in enumerate(raw_headers)
            if clean(header)
        }

    def rows(self):
        yield from self.worksheet.iter_rows(min_row=2, values_only=True)

    def value(self, row: tuple[Any, ...], *headers: str) -> Any:
        for header in headers:
            index = self.columns.get(normalize_header(header))
            if index is not None and index < len(row):
                value = row[index]
                if value is not None and clean(value):
                    return value
        return None

    def close(self) -> None:
        self.workbook.close()


class ExcelImporter:
    def __init__(self, directory: Path, apply: bool, overwrite: bool):
        self.directory = directory
        self.apply = apply
        self.overwrite = overwrite
        self.db: Session = SessionLocal()
        self.mappings: dict[tuple[str, str], int] = {}
        self.active_ids: dict[str, set[int]] = defaultdict(set)
        self.pending_primary_contacts: list[tuple[int, str, str]] = []
        self.stats: dict[str, Stats] = {}
        self.warnings: list[str] = []
        self._accounts_by_name: dict[str, list[models.Account]] | None = None
        self._contacts: list[models.Contact] | None = None
        self._vaults_by_name: dict[str, list[models.Vault]] | None = None
        self._activities_by_key: dict[tuple[str, datetime | None, datetime | None], list[models.Activity]] | None = None
        self._task_types_by_name: dict[str, models.TaskType] | None = None

    def load_mappings(self) -> None:
        try:
            rows = self.db.execute(
                text("SELECT entity_type, source_id, local_id FROM legacy_import_ids")
            )
        except Exception as exc:
            raise RuntimeError("Run python migrate_db.py before this importer.") from exc
        self.mappings = {(row[0], row[1]): row[2] for row in rows}

    def mapped(self, entity_type: str, source_id: str, model):
        local_id = self.mappings.get((entity_type, source_id))
        return self.db.get(model, local_id) if local_id else None

    def remember(self, entity_type: str, source_id: str, local_id: int) -> None:
        if not source_id:
            return
        self.mappings[(entity_type, source_id)] = local_id
        if not self.apply:
            return
        self.db.execute(
            text(
                "INSERT INTO legacy_import_ids (entity_type, source_id, local_id) "
                "VALUES (:entity_type, :source_id, :local_id) "
                "ON DUPLICATE KEY UPDATE local_id = VALUES(local_id)"
            ),
            {"entity_type": entity_type, "source_id": source_id, "local_id": local_id},
        )

    def set_value(self, obj, field: str, value: Any, overwrite: bool | None = None) -> bool:
        if value is None or value == "":
            return False
        current = getattr(obj, field, None)
        may_overwrite = self.overwrite if overwrite is None else overwrite
        if current not in (None, "") and not may_overwrite:
            return False
        if current == value:
            return False
        setattr(obj, field, value)
        return True

    def set_activity(self, obj, entity_type: str, active: bool) -> bool:
        if active:
            self.active_ids[entity_type].add(obj.id)
            if obj.is_active is not True:
                obj.is_active = True
                return True
            return False
        if obj.id in self.active_ids[entity_type]:
            return False
        if obj.is_active is not False:
            obj.is_active = False
            return True
        return False

    def workbook(self, key: str) -> WorkbookRows:
        path = self.directory / FILES[key]
        if not path.is_file():
            raise FileNotFoundError(f"Missing required export: {path}")
        return WorkbookRows(path)

    def import_accounts(self, key: str, active: bool) -> None:
        label = "Active Accounts" if active else "Inactive Accounts"
        stats = self.stats[label] = Stats()
        book = self.workbook(key)
        accounts = self.db.query(models.Account).all()
        by_name: dict[str, list[models.Account]] = defaultdict(list)
        for account in accounts:
            by_name[normalized(account.name)].append(account)

        try:
            for row in book.rows():
                stats.scanned += 1
                source_id = clean(book.value(row, "(Do Not Modify) Account"))
                name = clean(book.value(row, "Comapny Name", "Company Name", "Account Name"))
                if not source_id or not name:
                    stats.skipped += 1
                    continue

                account = self.mapped("account", source_id, models.Account)
                if account is None:
                    candidates = by_name.get(normalized(name), [])
                    account = candidates[0] if len(candidates) == 1 else None
                created = account is None
                if created:
                    account = models.Account(name=name, is_active=active)
                    self.db.add(account)
                    self.db.flush()
                    accounts.append(account)
                    by_name[normalized(name)].append(account)
                    stats.created += 1
                else:
                    stats.matched += 1

                changed = self.set_value(account, "name", name)
                changed |= self.set_value(account, "phone", clean(book.value(row, "Phone", "Main Phone")))
                changed |= self.set_value(account, "street", clean(book.value(row, "Street")))
                changed |= self.set_value(account, "zip_code", clean(book.value(row, "ZIP/Postal Code")))
                changed |= self.set_value(account, "city", clean(book.value(row, "City")))
                changed |= self.set_value(account, "country", clean(book.value(row, "Country/Region")))
                changed |= self.set_value(account, "website", clean(book.value(row, "Website")))
                changed |= self.set_activity(account, "account", active)
                self.remember("account", source_id, account.id)

                primary_name = clean(book.value(row, "Primary Contact"))
                primary_email = clean(book.value(row, "Email (Primary Contact) (Contact)"))
                if primary_name or primary_email:
                    self.pending_primary_contacts.append((account.id, primary_name, primary_email))
                if changed and not created:
                    stats.changed += 1
        finally:
            book.close()
        self._accounts_by_name = by_name

    def account_by_name(self, name: Any) -> models.Account | None:
        key = normalized(name)
        if not key:
            return None
        if self._accounts_by_name is None:
            self._accounts_by_name = defaultdict(list)
            for account in self.db.query(models.Account).all():
                self._accounts_by_name[normalized(account.name)].append(account)
        matches = self._accounts_by_name.get(key, [])
        return matches[0] if len(matches) == 1 else None

    def find_contact(
        self,
        source_id: str,
        email: str,
        first_name: str,
        last_name: str,
        account_id: int | None,
        phone: str,
    ) -> models.Contact | None:
        contact = self.mapped("contact", source_id, models.Contact)
        if contact:
            return contact
        if self._contacts is None:
            self._contacts = self.db.query(models.Contact).all()
        contacts = self._contacts
        if email:
            matches = [c for c in contacts if normalized(c.email) == normalized(email)]
            if len(matches) == 1:
                return matches[0]
        matches = [
            c for c in contacts
            if normalized(c.first_name) == normalized(first_name)
            and normalized(c.last_name) == normalized(last_name)
            and c.account_id == account_id
        ]
        if len(matches) == 1:
            return matches[0]
        phone_key = normalized_phone(phone)
        if phone_key:
            phone_matches = [c for c in matches if normalized_phone(c.phone) == phone_key]
            if len(phone_matches) == 1:
                return phone_matches[0]
        return None

    def import_contacts(self, key: str, active: bool) -> None:
        label = "Active Contacts" if active else "Inactive Contacts"
        stats = self.stats[label] = Stats()
        book = self.workbook(key)
        if self._contacts is None:
            self._contacts = self.db.query(models.Contact).all()
        try:
            for row in book.rows():
                stats.scanned += 1
                source_id = clean(book.value(row, "(Do Not Modify) Contact"))
                first, middle, last = split_name(
                    book.value(row, "First Name"),
                    book.value(row, "Middle Name"),
                    book.value(row, "Last Name"),
                    book.value(row, "Full Name"),
                )
                if not source_id or (not first and not last):
                    stats.skipped += 1
                    continue
                email = clean(book.value(row, "Email"))
                phone = clean(book.value(row, "Business Phone"))
                mobile = clean(book.value(row, "Mobile Phone"))
                account = self.account_by_name(book.value(row, "Account"))
                account_id = account.id if account else None
                if book.value(row, "Account") and account is None:
                    stats.unresolved += 1
                    self.warnings.append(
                        f"{label}: account '{clean(book.value(row, 'Account'))}' was not uniquely matched "
                        f"for contact '{clean(book.value(row, 'Full Name'))}'."
                    )

                contact = self.find_contact(
                    source_id, email, first, last, account_id, phone
                )
                created = contact is None
                if created:
                    israeli = parse_yes_no(book.value(row, "Israeli?"))
                    contact = models.Contact(
                        first_name=first,
                        last_name=last,
                        middle_name=middle or None,
                        email=email or None,
                        phone=phone or None,
                        mobile_phone=mobile or None,
                        is_israeli=israeli,
                        account_id=account_id,
                        is_active=active,
                    )
                    self.db.add(contact)
                    self.db.flush()
                    self._contacts.append(contact)
                    stats.created += 1
                else:
                    stats.matched += 1

                changed = self.set_value(contact, "first_name", first)
                changed |= self.set_value(contact, "middle_name", middle)
                changed |= self.set_value(contact, "last_name", last)
                changed |= self.set_value(contact, "email", email)
                changed |= self.set_value(contact, "phone", phone)
                changed |= self.set_value(contact, "mobile_phone", mobile)
                if account_id is not None:
                    changed |= self.set_value(contact, "account_id", account_id)
                israeli = parse_yes_no(book.value(row, "Israeli?"))
                if israeli is not None:
                    changed |= self.set_value(contact, "is_israeli", israeli)
                changed |= self.set_activity(contact, "contact", active)
                self.remember("contact", source_id, contact.id)
                if changed and not created:
                    stats.changed += 1
        finally:
            book.close()

    def link_primary_contacts(self) -> None:
        contacts = self._contacts if self._contacts is not None else self.db.query(models.Contact).all()
        stats = self.stats.setdefault("Primary Contacts", Stats())
        for account_id, name, email in self.pending_primary_contacts:
            stats.scanned += 1
            candidates = []
            if email:
                candidates = [c for c in contacts if normalized(c.email) == normalized(email)]
            if not candidates and name:
                candidates = [
                    c for c in contacts
                    if normalized(f"{c.first_name} {c.last_name}") == normalized(name)
                    and (c.account_id == account_id or c.account_id is None)
                ]
            if len(candidates) != 1:
                stats.unresolved += 1
                self.warnings.append(
                    f"Primary Contacts: '{name or email}' was not uniquely matched for account ID {account_id}."
                )
                continue
            account = self.db.get(models.Account, account_id)
            if account and self.set_value(account, "primary_contact_id", candidates[0].id):
                stats.changed += 1
            else:
                stats.matched += 1

    def import_contracts(self, key: str, active: bool) -> None:
        label = "Active Contracts" if active else "Inactive Contracts"
        stats = self.stats[label] = Stats()
        book = self.workbook(key)
        contracts = self.db.query(models.Contract).all()
        by_title: dict[str, list[models.Contract]] = defaultdict(list)
        for contract in contracts:
            by_title[normalized(contract.title)].append(contract)
        try:
            for row in book.rows():
                stats.scanned += 1
                source_id = clean(book.value(row, "(Do Not Modify) Contract"))
                title = clean(book.value(row, "Name", "Contract Title"))
                if not source_id or not title:
                    stats.skipped += 1
                    continue
                contract = self.mapped("contract", source_id, models.Contract)
                if contract is None:
                    candidates = by_title.get(normalized(title), [])
                    contract = candidates[0] if len(candidates) == 1 else None
                beneficiary, supplier = split_contract_title(title)
                created = contract is None
                if created:
                    contract = models.Contract(
                        title=title,
                        beneficiary_title=beneficiary or None,
                        supplier_title=supplier or None,
                        contact_type="3-party",
                        status=models.ContractStatus.ACTIVE,
                        value=0.0,
                        is_active=active,
                    )
                    self.db.add(contract)
                    self.db.flush()
                    contracts.append(contract)
                    by_title[normalized(title)].append(contract)
                    stats.created += 1
                else:
                    stats.matched += 1

                changed = self.set_value(contract, "title", title)
                changed |= self.set_value(contract, "beneficiary_title", beneficiary)
                changed |= self.set_value(contract, "supplier_title", supplier)
                changed |= self.set_value(contract, "start_date", parse_date(book.value(row, "Date Contract Signed")))
                changed |= self.set_value(contract, "end_date", parse_date(book.value(row, "Date Contract Ends")))
                changed |= self.set_value(
                    contract,
                    "beneficiary_management_contact",
                    clean(book.value(row, "B Management Contact", "Beneficiary Management Contact")),
                )
                changed |= self.set_value(
                    contract,
                    "supplier_management_contact",
                    clean(book.value(row, "S Management Contact", "Supplier Management Contact")),
                )
                changed |= self.set_activity(contract, "contract", active)
                self.remember("contract", source_id, contract.id)
                if changed and not created:
                    stats.changed += 1
        finally:
            book.close()

    def vault_by_name(self, name: Any) -> models.Vault | None:
        key = normalized(name)
        if not key:
            return None
        if self._vaults_by_name is None:
            self._vaults_by_name = defaultdict(list)
            for vault in self.db.query(models.Vault).all():
                self._vaults_by_name[normalized(vault.name)].append(vault)
        matches = self._vaults_by_name.get(key, [])
        if len(matches) == 1:
            return matches[0]
        if not matches:
            vault = models.Vault(name=clean(name), status=models.VaultStatus.OPEN, is_active=True)
            self.db.add(vault)
            self.db.flush()
            self._vaults_by_name[key].append(vault)
            return vault
        return None

    def import_deposits(self, key: str, active: bool) -> None:
        label = "Active Deposits" if active else "Inactive Deposits"
        stats = self.stats[label] = Stats()
        book = self.workbook(key)
        try:
            for row in book.rows():
                stats.scanned += 1
                source_id = clean(book.value(row, "(Do Not Modify) Deposit"))
                if not source_id:
                    stats.skipped += 1
                    continue
                reference = clean(book.value(row, "Deposit Number"))
                deposit = self.mapped("deposit", source_id, models.Deposit)
                if deposit is None and reference:
                    deposit = self.db.query(models.Deposit).filter(
                        models.Deposit.reference_number == reference
                    ).first()
                created = deposit is None
                if created:
                    safe_reference = reference or f"IMPORT-{source_id}"
                    deposit = models.Deposit(
                        reference_number=safe_reference,
                        amount=0.0,
                        status=models.DepositStatus.CLEARED,
                        is_active=active,
                    )
                    self.db.add(deposit)
                    self.db.flush()
                    stats.created += 1
                else:
                    stats.matched += 1

                changed = self.set_value(deposit, "reference_number", reference)
                changed |= self.set_value(deposit, "product_name", clean(book.value(row, "Product Name")))
                changed |= self.set_value(deposit, "version", clean(book.value(row, "Version")))
                changed |= self.set_value(deposit, "supplier", clean(book.value(row, "Supplier")))
                changed |= self.set_value(deposit, "date", parse_date(book.value(row, "Date Received")))
                changed |= self.set_value(deposit, "box", clean(book.value(row, "Box")))
                changed |= self.set_value(deposit, "received_by", clean(book.value(row, "Received By")))
                vault = self.vault_by_name(book.value(row, "Vault"))
                if vault:
                    changed |= self.set_value(deposit, "vault_id", vault.id)
                changed |= self.set_activity(deposit, "deposit", active)
                self.remember("deposit", source_id, deposit.id)
                if changed and not created:
                    stats.changed += 1
        finally:
            book.close()

    def task_type(self, name: str) -> models.TaskType:
        if self._task_types_by_name is None:
            self._task_types_by_name = {
                normalized(task_type.name): task_type
                for task_type in self.db.query(models.TaskType).all()
            }
        key = normalized(name)
        task_type = self._task_types_by_name.get(key)
        if task_type:
            return task_type
        task_type = models.TaskType(name=name[:50], color="#6366f1")
        self.db.add(task_type)
        self.db.flush()
        self._task_types_by_name[key] = task_type
        return task_type

    def import_activities(self) -> None:
        stats = self.stats["Activities"] = Stats()
        book = self.workbook("activities")
        if self._activities_by_key is None:
            self._activities_by_key = defaultdict(list)
            for existing in self.db.query(models.Activity).all():
                self._activities_by_key[
                    (normalized(existing.subject), existing.start_date, existing.due_date)
                ].append(existing)
        try:
            for row in book.rows():
                stats.scanned += 1
                source_id = clean(book.value(row, "(Do Not Modify) Activity"))
                subject = clean(book.value(row, "Subject"))
                if not source_id or not subject:
                    stats.skipped += 1
                    continue
                activity = self.mapped("activity", source_id, models.Activity)
                if activity is None:
                    start = parse_date(book.value(row, "Start Date"))
                    due = parse_date(book.value(row, "Due Date"))
                    candidates = self._activities_by_key.get((normalized(subject), start, due), [])
                    activity = candidates[0] if len(candidates) == 1 else None
                created = activity is None
                if created:
                    activity = models.Activity(subject=subject, is_active=True)
                    self.db.add(activity)
                    self.db.flush()
                    self._activities_by_key[
                        (
                            normalized(subject),
                            parse_date(book.value(row, "Start Date")),
                            parse_date(book.value(row, "Due Date")),
                        )
                    ].append(activity)
                    stats.created += 1
                else:
                    stats.matched += 1
                activity_type = (
                    activity_type_from_subject(subject)
                    or clean(book.value(row, "Activity Type"))
                    or "Task"
                )
                task_type = self.task_type(activity_type)
                changed = self.set_value(activity, "subject", subject)
                changed |= self.set_value(activity, "regarding", clean(book.value(row, "Regarding")))
                changed |= self.set_value(activity, "start_date", parse_date(book.value(row, "Start Date")))
                changed |= self.set_value(activity, "due_date", parse_date(book.value(row, "Due Date")))
                changed |= self.set_value(activity, "task_type_id", task_type.id)
                changed |= self.set_activity(activity, "activity", True)
                self.remember("activity", source_id, activity.id)
                if changed and not created:
                    stats.changed += 1
        finally:
            book.close()

    def run(self) -> None:
        self.load_mappings()
        # Active rows go first so they win if an export contains an overlap.
        self.import_accounts("active_accounts", True)
        self.import_accounts("inactive_accounts", False)
        self.import_contacts("active_contacts", True)
        self.import_contacts("inactive_contacts", False)
        self.link_primary_contacts()
        self.import_contracts("active_contracts", True)
        self.import_contracts("inactive_contracts", False)
        self.import_deposits("active_deposits", True)
        self.import_deposits("inactive_deposits", False)
        self.import_activities()

        if self.apply:
            self.db.commit()
        else:
            self.db.rollback()

        print("\nImport Summary")
        for label, stats in self.stats.items():
            stats.display(label)
        if self.warnings:
            print("\nUnresolved relationships")
            for warning in self.warnings:
                print(f"  - {warning}")
        print("Changes committed." if self.apply else "Dry run only; no imported data was committed.")

    def close(self) -> None:
        self.db.close()


def validate_files(directory: Path) -> list[str]:
    return [name for name in FILES.values() if not (directory / name).is_file()]


def parse_args():
    parser = argparse.ArgumentParser(description="Import active and inactive EGCRM Excel exports.")
    parser.add_argument("directory", type=Path, help="Directory containing the nine .xlsx exports.")
    parser.add_argument("--apply", action="store_true", help="Commit changes; otherwise run a dry run.")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace populated CRM fields with non-empty spreadsheet values.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    directory = args.directory.expanduser().resolve()
    missing = validate_files(directory)
    if missing:
        print("Missing required exports:")
        for name in missing:
            print(f"  - {name}")
        return 2
    print(f"Export directory: {directory}")
    print(f"Mode: {'APPLY' if args.apply else 'DRY RUN'}")
    print(f"Overwrite populated fields: {'yes' if args.overwrite else 'no'}")
    print(
        "Coverage note: Role is absent from both Contacts exports; inactive Contacts also "
        "omits email, mobile phone, and Israeli status. Contract exports provide only the "
        "beneficiary and supplier management contacts in the inactive file. Existing CRM "
        "values in absent columns will be preserved."
    )
    importer = ExcelImporter(directory, args.apply, args.overwrite)
    try:
        importer.run()
    except Exception:
        importer.db.rollback()
        raise
    finally:
        importer.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
