"""
EGCRM - Unified Data Import Script
Imports all data from Excel exports in the correct dependency order:
  1. Accounts (must be first — contacts and contracts depend on them)
  2. Contacts (linked to accounts)
  3. Leads (linked to contacts)
  4. Vaults
  5. Contracts (linked to accounts)
  6. Deposits (linked to vaults)

Usage:
    cd /home/egcrm/backend
    source venv/bin/activate
    python import_all.py
"""
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from datetime import datetime
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import SessionLocal
import models

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'excel_exports')

STATUS_MAP = {
    'New': 'New',
    'Contacted': 'Contacted',
    'Qualified': 'Qualified',
    'Lost': 'Lost',
    'Cannot Contact': 'Lost',
    'No Longer Interested': 'Lost',
    'Canceled': 'Lost',
}

def str_val(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M:%S")
    except:
        try:
            return datetime.strptime(str(val).strip(), "%d/%m/%Y")
        except:
            return None


# ── 1. Accounts ───────────────────────────────────────────────────────────────
def import_accounts(db: Session):
    print("\n--- Importing Accounts ---")
    path = os.path.join(EXPORTS_DIR, 'Active Accounts 26-Feb-26 1-28-48 PM.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h}

    imported = updated = skipped = 0
    account_map = {}  # name -> id, returned so contacts can use it

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str_val(row[col.get('Account Name', col.get('Name', 0))])
        if not name:
            skipped += 1
            continue

        phone    = str_val(row[col.get('Main Phone',    col.get('Phone', -1))]) if col.get('Main Phone', col.get('Phone')) is not None else None
        website  = str_val(row[col.get('Website',       -1)]) if 'Website'       in col else None
        industry = str_val(row[col.get('Industry',      -1)]) if 'Industry'      in col else None
        street   = str_val(row[col.get('Address 1: Street 1', col.get('Street', -1))]) if col.get('Address 1: Street 1', col.get('Street')) is not None else None
        city     = str_val(row[col.get('Address 1: City', col.get('City', -1))]) if col.get('Address 1: City', col.get('City')) is not None else None
        country  = str_val(row[col.get('Address 1: Country/Region', col.get('Country', -1))]) if col.get('Address 1: Country/Region', col.get('Country')) is not None else None

        existing = db.query(models.Account).filter(models.Account.name == name).first()
        if existing:
            account_map[name.lower()] = existing.id
            # Update empty fields
            if phone    and not existing.phone:    existing.phone    = phone
            if website  and not existing.website:  existing.website  = website
            if industry and not existing.industry: existing.industry = industry
            if street   and not existing.street:   existing.street   = street
            if city     and not existing.city:     existing.city     = city
            if country  and not existing.country:  existing.country  = country
            db.commit()
            updated += 1
            continue

        acc = models.Account(
            name=name,
            phone=phone,
            website=website,
            industry=industry,
            street=street,
            city=city,
            country=country,
        )
        db.add(acc)
        db.flush()
        account_map[name.lower()] = acc.id
        db.commit()
        imported += 1

    wb.close()
    print(f"  Accounts: imported={imported}, updated={updated}, skipped={skipped}")
    return account_map


# ── 2. Contacts ───────────────────────────────────────────────────────────────
def import_contacts(db: Session, account_map: dict):
    print("\n--- Importing Contacts ---")
    path = os.path.join(EXPORTS_DIR, 'My Active Contacts 26-Feb-26 1-29-17 PM.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    imported = updated = skipped = 0
    seen_emails = set()
    seen_names  = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        first_name = str_val(row[col.get('First Name', 4)])
        last_name  = str_val(row[col.get('Last Name',  6)])
        if not first_name and not last_name:
            skipped += 1
            continue

        first_name = first_name or ''
        last_name  = last_name  or ''

        email       = str_val(row[col.get('Email',          7)])
        middle_name = str_val(row[col.get('Middle Name',    5)])
        phone       = str_val(row[col.get('Business Phone', 9)])
        mobile      = str_val(row[col.get('Mobile Phone',  10)])
        account_raw = str_val(row[col.get('Account',        8)])
        israeli_raw = str_val(row[col.get('Israeli?',      11)])
        is_israeli  = True if israeli_raw and israeli_raw.lower() == 'yes' else False

        account_id = None
        if account_raw:
            account_id = account_map.get(account_raw.strip().lower())

        if email:
            email_lower = email.lower()
            if email_lower in seen_emails:
                skipped += 1
                continue
            seen_emails.add(email_lower)

            existing = db.query(models.Contact).filter(models.Contact.email == email).first()
            if existing:
                existing.middle_name  = middle_name
                existing.mobile_phone = mobile
                existing.is_israeli   = is_israeli
                if account_id and not existing.account_id:
                    existing.account_id = account_id
                db.commit()
                updated += 1
                continue
        else:
            name_key = (first_name.lower(), last_name.lower())
            if name_key in seen_names:
                skipped += 1
                continue
            seen_names.add(name_key)

            existing = db.query(models.Contact).filter(
                models.Contact.first_name == first_name,
                models.Contact.last_name  == last_name,
                models.Contact.email      == None,
            ).first()
            if existing:
                existing.middle_name  = middle_name
                existing.mobile_phone = mobile
                existing.is_israeli   = is_israeli
                db.commit()
                updated += 1
                continue

        db.add(models.Contact(
            first_name=first_name, last_name=last_name, middle_name=middle_name,
            email=email, phone=phone, mobile_phone=mobile,
            is_israeli=is_israeli, account_id=account_id,
        ))
        db.commit()
        imported += 1

    wb.close()
    print(f"  Contacts: imported={imported}, updated={updated}, skipped={skipped}")


# ── 3. Leads ──────────────────────────────────────────────────────────────────
def import_leads(db: Session):
    print("\n--- Importing Leads ---")
    path = os.path.join(EXPORTS_DIR, 'My Open Leads 26-Feb-26 4-36-09 PM.xlsx')
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    existing_titles = {
        r[0].lower() for r in db.execute(text("SELECT title FROM leads")).fetchall()
    }
    contacts_raw = db.execute(
        text("SELECT id, LOWER(first_name), LOWER(last_name) FROM contacts")
    ).fetchall()
    contact_map = {(r[1].strip(), r[2].strip()): r[0] for r in contacts_raw}

    imported = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        first      = str_val(row[col.get('First Name',    0)]) or ''
        last       = str_val(row[col.get('Last Name',     2)]) or ''
        topic      = str_val(row[col.get('Topic',         3)]) or ''
        status_raw = str_val(row[col.get('Status Reason', 4)]) or 'New'

        title = topic if topic else f"{first} {last}".strip()
        if not title:
            skipped += 1
            continue

        if title.lower() in existing_titles:
            skipped += 1
            continue

        status = STATUS_MAP.get(status_raw, 'New')
        contact_id = contact_map.get((first.lower(), last.lower()))

        db.execute(
            text("INSERT INTO leads (title, status, value, contact_id, assigned_to_user_id, created_at) VALUES (:title, :status, 0.0, :cid, NULL, NOW())"),
            {"title": title, "status": status, "cid": contact_id}
        )
        existing_titles.add(title.lower())
        imported += 1

    db.commit()
    wb.close()
    print(f"  Leads: imported={imported}, skipped={skipped}")


# ── 4. Vaults ─────────────────────────────────────────────────────────────────
def import_vaults(db: Session):
    print("\n--- Importing Vaults ---")
    path = os.path.join(EXPORTS_DIR, 'Active Vaults 26-Feb-26 1-29-36 PM.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    imported = skipped = 0
    vault_map = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str_val(row[col.get('Vault Name', 3)])
        if not name:
            skipped += 1
            continue

        existing = db.query(models.Vault).filter(models.Vault.name == name).first()
        if existing:
            vault_map[name] = existing.id
            skipped += 1
            continue

        v = models.Vault(name=name, status=models.VaultStatus.OPEN)
        db.add(v)
        db.flush()
        vault_map[name] = v.id
        imported += 1

    db.commit()
    wb.close()
    print(f"  Vaults: imported={imported}, skipped={skipped}")
    return vault_map


# ── 5. Contracts ──────────────────────────────────────────────────────────────
def import_contracts(db: Session):
    print("\n--- Importing Contracts ---")
    path = os.path.join(EXPORTS_DIR, 'Active Contracts 26-Feb-26 1-27-18 PM.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    imported = skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = str_val(row[col.get('Name', 3)])
        if not title:
            skipped += 1
            continue

        if db.query(models.Contract).filter(models.Contract.title == title).first():
            skipped += 1
            continue

        db.add(models.Contract(
            title=title,
            status=models.ContractStatus.ACTIVE,
            start_date=parse_date(row[col.get('Date Contract Signed', 4)]),
            end_date=parse_date(row[col.get('Date Contract Ends', 5)]),
            value=0.0,
        ))
        imported += 1

    db.commit()
    wb.close()
    print(f"  Contracts: imported={imported}, skipped={skipped}")


# ── 6. Deposits ───────────────────────────────────────────────────────────────
def import_deposits(db: Session, vault_map: dict):
    print("\n--- Importing Deposits ---")
    path = os.path.join(EXPORTS_DIR, 'Active Deposits 26-Feb-26 1-29-51 PM.xlsx')
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    imported = updated = skipped = 0
    seen_refs = set()

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        deposit_number = row[col.get('Deposit Number', 9)]
        ref = str(deposit_number).strip() if deposit_number else f"DEP-{row_num}"
        if ref in seen_refs:
            skipped += 1
            continue
        seen_refs.add(ref)

        vault_name = str_val(row[col.get('Vault', 7)])
        vault_id = vault_map.get(vault_name) if vault_name else None

        existing = db.query(models.Deposit).filter(models.Deposit.reference_number == ref).first()
        if existing:
            if vault_id and not existing.vault_id:
                existing.vault_id = vault_id
            db.commit()
            updated += 1
            continue

        db.add(models.Deposit(
            reference_number=ref,
            amount=0.0,
            date=parse_date(row[col.get('Date Received', 6)]),
            status=models.DepositStatus.CLEARED,
            vault_id=vault_id,
            product_name=str_val(row[col.get('Product Name', 3)]),
            version=str_val(row[col.get('Version', 4)]),
            supplier=str_val(row[col.get('Supplier', 5)]),
            box=str_val(row[col.get('Box', 8)]),
            received_by=str_val(row[col.get('Received By', 10)]),
        ))
        db.commit()
        imported += 1

    wb.close()
    print(f"  Deposits: imported={imported}, updated={updated}, skipped={skipped}")


# ── Main ──────────────────────────────────────────────────────────────────────
def run():
    db: Session = SessionLocal()
    try:
        print("=" * 50)
        print("  EGCRM Full Data Import")
        print("=" * 50)

        account_map = import_accounts(db)   # 1. Accounts first
        import_contacts(db, account_map)    # 2. Contacts (needs account_map)
        import_leads(db)                    # 3. Leads
        vault_map = import_vaults(db)       # 4. Vaults
        import_contracts(db)               # 5. Contracts
        import_deposits(db, vault_map)      # 6. Deposits (needs vault_map)

        print("\n" + "=" * 50)
        print("  ✅ All imports complete!")
        print("=" * 50)
    except Exception as e:
        db.rollback()
        print(f"\n❌ ERROR: {e}")
        import traceback; traceback.print_exc()
    finally:
        db.close()


if __name__ == '__main__':
    run()
