"""
Import script for:
- Active Contracts
- Active Vaults
- Active Deposits
"""
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.dirname(__file__))

import openpyxl
from datetime import datetime
from sqlalchemy.orm import Session
from database import SessionLocal
import models

EXPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'excel_exports')

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d %H:%M:%S")
    except:
        try:
            return datetime.strptime(str(val).strip(), "%d/%m/%Y")
        except:
            return None

def import_vaults(db: Session):
    print("\n--- Importing Vaults ---")
    wb = openpyxl.load_workbook(os.path.join(EXPORTS_DIR, 'Active Vaults 26-Feb-26 1-29-36 PM.xlsx'))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    imported = skipped = 0
    vault_name_to_id = {}  # for Deposits lookup later

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[col.get('Vault Name', 3)]
        if not name:
            skipped += 1
            continue
        name = str(name).strip()
        exists = db.query(models.Vault).filter(models.Vault.name == name).first()
        if exists:
            vault_name_to_id[name] = exists.id
            skipped += 1
            continue
        v = models.Vault(name=name, status=models.VaultStatus.OPEN)
        db.add(v)
        db.flush()  # get id
        vault_name_to_id[name] = v.id
        imported += 1

    db.commit()
    print(f"  Vaults: imported={imported}, skipped={skipped}")
    return vault_name_to_id

def import_contracts(db: Session):
    print("\n--- Importing Contracts ---")
    wb = openpyxl.load_workbook(os.path.join(EXPORTS_DIR, 'Active Contracts 26-Feb-26 1-27-18 PM.xlsx'))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    imported = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[col.get('Name', 3)]
        if not title:
            skipped += 1
            continue
        title = str(title).strip()

        exists = db.query(models.Contract).filter(models.Contract.title == title).first()
        if exists:
            skipped += 1
            continue

        start_date = parse_date(row[col.get('Date Contract Signed', 4)])
        end_date = parse_date(row[col.get('Date Contract Ends', 5)])

        c = models.Contract(
            title=title,
            status=models.ContractStatus.ACTIVE,
            start_date=start_date,
            end_date=end_date,
            value=0.0,
        )
        db.add(c)
        imported += 1

    db.commit()
    print(f"  Contracts: imported={imported}, skipped={skipped}")

def import_deposits(db: Session, vault_name_to_id: dict):
    print("\n--- Importing Deposits ---")
    wb = openpyxl.load_workbook(os.path.join(EXPORTS_DIR, 'Active Deposits 26-Feb-26 1-29-51 PM.xlsx'))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    def sv(val):
        if val is None: return None
        s = str(val).strip()
        return s if s else None

    imported = updated = skipped = 0
    seen_refs = set()  # handle duplicates within the Excel itself

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        deposit_number = row[col.get('Deposit Number', 9)]
        ref = str(deposit_number).strip() if deposit_number else f"DEP-{row_num}"
        if not ref:
            ref = f"DEP-{row_num}"

        # Skip within-file duplicates (same ref seen more than once in Excel)
        if ref in seen_refs:
            skipped += 1
            continue
        seen_refs.add(ref)

        vault_name = row[col.get('Vault', 7)]
        vault_id = None
        if vault_name:
            vault_id = vault_name_to_id.get(str(vault_name).strip())

        date_received = parse_date(row[col.get('Date Received', 6)])
        pname       = sv(row[col.get('Product Name', 3)])
        version     = sv(row[col.get('Version', 4)])
        supplier    = sv(row[col.get('Supplier', 5)])
        box         = sv(row[col.get('Box', 8)])
        received_by = sv(row[col.get('Received By', 10)])

        exists = db.query(models.Deposit).filter(models.Deposit.reference_number == ref).first()
        if exists:
            exists.product_name = pname
            exists.version      = version
            exists.supplier     = supplier
            exists.box          = box
            exists.received_by  = received_by
            if vault_id and not exists.vault_id:
                exists.vault_id = vault_id
            db.commit()
            updated += 1
            continue

        d = models.Deposit(
            reference_number=ref,
            amount=0.0,
            date=date_received,
            status=models.DepositStatus.CLEARED,
            vault_id=vault_id,
            product_name=pname,
            version=version,
            supplier=supplier,
            box=box,
            received_by=received_by,
        )
        db.add(d)
        db.commit()
        imported += 1

    print(f"  Deposits: imported={imported}, updated={updated}, skipped={skipped}")

def run():
    db: Session = SessionLocal()
    try:
        vault_map = import_vaults(db)
        import_contracts(db)
        import_deposits(db, vault_map)
        print("\nAll imports complete.")
    except Exception as e:
        db.rollback()
        print(f"ERROR: {e}")
        raise
    finally:
        db.close()

if __name__ == '__main__':
    run()
